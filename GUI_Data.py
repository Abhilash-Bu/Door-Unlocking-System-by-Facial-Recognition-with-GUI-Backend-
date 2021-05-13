from openpyxl import *
from tkinter import *


def gui_storing():
    wb = load_workbook('Data_of_Employees.xlsx')
    sheet = wb.active

    def excel():
        # resize the width of columns in excel spreadsheet
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30

        # write given data to an excel spreadsheet at particular location
        sheet.cell(row=1, column=1).value = "Employee_ID"
        sheet.cell(row=1, column=2).value = "Name"
        sheet.cell(row=1, column=3).value = "Phone_Number"
        sheet.cell(row=1, column=4).value = "Twilio_Number"
        sheet.cell(row=1, column=5).value = "Twilio_Token"
        sheet.cell(row=1, column=6).value = "Twilio_SSID"

    def insert_user():
        current_row = sheet.max_row
        global emp_value
        emp_value = sheet.cell(row=current_row, column=1).value +1
        sheet.cell(row=current_row + 1, column=1).value = emp_value
        sheet.cell(row=current_row + 1, column=2).value = name_field.get()
        sheet.cell(row=current_row + 1, column=3).value = phoneno_field.get()

        wb.save('Data_of_Employees.xlsx')

        name_field.delete(0, END)
        phoneno_field.delete(0, END)
        root.destroy()

    def insert_admin():
        current_row = sheet.max_row
        global emp_value
        emp_value = 1
        sheet.cell(row=current_row + 1, column=1).value = emp_value
        sheet.cell(row=current_row + 1, column=2).value = name_field.get()
        sheet.cell(row=current_row + 1, column=3).value = phoneno_field.get()
        sheet.cell(row=current_row + 1, column=4).value = tw_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = tw_token_field.get()
        sheet.cell(row=current_row + 1, column=6).value = tw_ssid_field.get()

        wb.save('Data_of_Employees.xlsx')

        name_field.delete(0, END)
        phoneno_field.delete(0, END)
        tw_no_field.delete(0, END)
        tw_token_field.delete(0, END)
        tw_ssid_field.delete(0, END)
        root.destroy()

    current_row = sheet.max_row
    if current_row == 1:  # ADMIN STORAGE
        root = Tk()
        root.configure(background='light green')
        root.title("Admin Form")
        root.geometry("600x300")
        excel()

        heading = Label(root, text="Admin Form (Create Twilio Account)", bg="light green", font='arial 15 bold')
        name = Label(root, text="Full Name", bg="light green")
        ph = Label(root, text="Phone Number(with country code)", bg="light green")
        tw_no = Label(root, text="Twilio Ph No", bg="light green")
        tw_token = Label(root, text="Twilio Token", bg="light green")
        tw_ssid = Label(root, text="Twilio SSID", bg="light green")

        heading.grid(row=0, column=1)
        name.grid(row=2, column=0)
        ph.grid(row=3, column=0)
        tw_no.grid(row=4, column=0)
        tw_token.grid(row=5, column=0)
        tw_ssid.grid(row=6, column=0)

        # create a text entry box for tying information
        name_field = Entry(root, textvariable = StringVar)
        phoneno_field = Entry(root, textvariable = StringVar)
        tw_no_field = Entry(root, textvariable = StringVar)
        tw_token_field = Entry(root, textvariable = StringVar)
        tw_ssid_field = Entry(root, textvariable = StringVar)

        # grid method is used for placing the widgets at respective positions
        name_field.grid(row=2, column=1, ipadx="100")
        phoneno_field.grid(row=3, column=1, ipadx="100")
        tw_no_field.grid(row=4, column=1, ipadx="100")
        tw_token_field.grid(row=5, column=1, ipadx="100")
        tw_ssid_field.grid(row=6, column=1, ipadx="100")

        excel()
        # create a Submit Button and place into the root window
        submit = Button(root, text="Submit", fg="Black",
                        bg="Red", command=insert_admin)
        submit.grid(row=11, column=1)
        # start the GUI
        root.mainloop()
        return emp_value

    else:  # USER STORAGE
        root = Tk()
        root.configure(background='light green')
        root.title("Registration Form")
        root.geometry("500x300")
        excel()
        heading = Label(root, text="User Form", bg="light green", font='arial 15 bold')
        name = Label(root, text="Full Name", bg="light green")
        ph = Label(root, text="Phone Number(with country code)", bg="light green")

        heading.grid(row=0, column=1)
        name.grid(row=2, column=0)
        ph.grid(row=3, column=0)

        name_field = Entry(root)
        phoneno_field = Entry(root)

        name_field.grid(row=2, column=1, ipadx="100")
        phoneno_field.grid(row=3, column=1, ipadx="100")

        excel()
        submit = Button(root, text="Submit", fg="Black",
                        bg="Red", command=insert_user)
        submit.grid(row=8, column=1)
        root.mainloop()

        master = Tk()
        master.configure(background='light green')
        master.geometry("500x300")
        master.title("Note")
        Label(master, text="Your Employee ID is : {}\nRemember for further use".format(emp_value),
              bg="light green", font='arial 15 bold').grid(row=0, column=1)

        Button(master, text="Quit", command=master.destroy).grid(row=3, column=1)
        return emp_value
